import pyodbc
import time
import os
from openpyxl import load_workbook
# from openpyxl.utils.cell import column_index_from_string
# config_column_email = column_index_from_string(config['mandatory_columns']['email'])
import configparser
import sys
import datetime
from datetime import datetime as dt

####################
# CONFIG: DATABASE #
####################

# DEVELOPMENT
db_credentials = ("Driver={SQL Server};" +
                  "Server=(local);" +
                  "Database=Zendesk;" +
                  "UID=zendesk_csp;" +
                  "PWD=xUG3y=4zp<sr?f>/;")

connection = pyodbc.connect(db_credentials)
cursor = connection.cursor()


######################
# FUNCTIONS: UTILITY #
######################

def timestamp():
    return "[" + time.strftime("%H:%M:%S", time.localtime()) + "] "


def ask(msg):
    return input(timestamp() + msg)


def out(msg):
    print(timestamp() + msg)


##########################
# FUNCTIONS: SQL QUERIES #
##########################

def sql_get_user_id(email):
    user_id = cursor.execute("select id from Users where email like '" + email + "'").fetchval()
    # user_id = None if no results
    return user_id


def sql_get_last_id(table):
    last_id = cursor.execute("select max(id) from " + table).fetchval()
    # last_id = None if no results
    if last_id is None:
        return 0
    return last_id


def sql_select_top(table):
    return cursor.execute("select top 1 * from " + table).fetchall()


###############################
# FUNCTIONS: DATABASE INSERTS #
###############################

def db_insert_survey(file):
    file_name = file[:len(file) - 5]  # remove .xlsx extension
    survey_type = get_survey_type()
    sql = "insert into CSP_Survey(name, upload_date, type_annual) values (?,getDate(),?)"
    cursor.execute(sql, (file_name, repr(survey_type)))
    connection.commit()
    return sql_get_last_id("CSP_Survey")


def db_insert_survey_user(survey_id, user_id, email, end_date):
    # email does not exist in Users table
    if user_id is None:
        # user_id will be NULL
        sql = "insert into CSP_Survey_User(survey_id, email, end_date) values (?,?,?)"
        cursor.execute(sql, (repr(survey_id), email, end_date))
        connection.commit()
        return sql_get_last_id("CSP_Survey_User")

    # email exists in Users table
    sql = "insert into CSP_Survey_User(survey_id, user_id, email, end_date) values (?,?,?,?)"
    cursor.execute(sql, (repr(survey_id), repr(user_id), email, end_date))
    connection.commit()
    return sql_get_last_id("CSP_Survey_User")


def db_insert_response(survey_user_id, question_id, score, comment):
    if score is None:
        sql = "insert into CSP_Response(survey_user_id, question_id, comment) values (?,?,?)"
        cursor.execute(sql, (repr(survey_user_id), repr(question_id), comment))
        connection.commit()
        return sql_get_last_id("CSP_Response")
    if comment is None:
        sql = "insert into CSP_Response(survey_user_id, question_id, score) values (?,?,?)"
        cursor.execute(sql, (repr(survey_user_id), repr(question_id), repr(score)))
        connection.commit()
        return sql_get_last_id("CSP_Response")
    # neither score nor comment are None
    sql = "insert into CSP_Response(survey_user_id, question_id, score, comment) values (?,?,?,?)"
    cursor.execute(sql, (repr(survey_user_id), repr(question_id), repr(score), comment))
    connection.commit()
    return sql_get_last_id("CSP_Response")


####################
# FUNCTIONS: OTHER #
####################

def check_config(config):
    # check if config file exists
    try:
        f = open('spreadsheet_config.ini')
    except IOError:
        out('Error: spreadsheet_config.ini: cannot open file\n'
            + 'Are you running process.py from the directory containing spreadsheet_config.ini?\n'
            + 'Please check that spreadsheet_config.ini exists and is named correctly')
        sys.exit(1)
    f.close()

    # check if section exists
    if 'mandatory_columns' not in config:
        out('Error: spreadsheet_config.ini: missing section [mandatory_columns]')
        sys.exit(1)

    # check if keys in section exist
    if 'email' not in config['mandatory_columns']:
        out('Error: spreadsheet_config.ini: missing key "email" in section [mandatory_columns]')
        sys.exit(1)
    if 'end_date' not in config['mandatory_columns']:
        out('Error: spreadsheet_config.ini: missing key "end_date" in section [mandatory_columns]')
        sys.exit(1)
    if 'end_date_format' not in config['mandatory_columns']:
        out('Error: spreadsheet_config.ini: missing key "end_date_format" in section [mandatory_columns]')
        sys.exit(1)

    email = config['mandatory_columns']['email']
    end_date = config['mandatory_columns']['end_date']
    end_date_format = config['mandatory_columns']['end_date_format']

    # check if keys have values
    if email == '':
        out('Error: spreadsheet_config.ini: no value for key "email" in section [mandatory_columns]')
        sys.exit(1)
    if end_date == '':
        out('Error: spreadsheet_config.ini: no value for key "end_date" in section [mandatory_columns]')
        sys.exit(1)
    if end_date_format == '':
        out('Error: spreadsheet_config.ini: no value for key "end_date_format" in section [mandatory_columns]')
        sys.exit(1)

    # check if key values are alphabetic only
    if not email.isalpha():
        out('Error: spreadsheet_config.ini: non-alphabetic value for key "email" in section [mandatory_columns]\n'
            + 'Value for "email": ' + email + '\n'
            + 'Please enter alphabetic-only values e.g. "AF" to indicate email column')
        sys.exit(1)
    if not end_date.isalpha():
        out('Error: spreadsheet_config.ini: non-alphabetic value for key "end_date" in section [mandatory_columns]\n'
            + 'Value for "end_date": ' + end_date + '\n'
            + 'Please enter alphabetic-only values e.g. "AF" to indicate end_date column')
        sys.exit(1)

    # check if end_date_format is valid
    if not (end_date_format == 'UK' or end_date_format == 'US'):
        out('Error: spreadsheet_config.ini: invalid value for key "end_date_format" in section [mandatory_columns]\n'
            + 'Value for "end_date_format": ' + end_date_format + '\n'
            + 'Please enter either "UK" or "US"')
        sys.exit(1)

    # config file is fine
    return


def get_path():
    user_input = ask("Enter directory path for survey results files: ")
    if os.path.exists(user_input):
        return user_input
    else:
        out("Invalid path: " + user_input)
        return get_path()


def get_file(path):
    print()
    directory_list = os.listdir(path)
    for i in range(len(directory_list)):
        print(" " + str(i) + " : " + directory_list[i])
    print()
    user_input = ask("Enter number for file to process: ")
    try:
        if 0 <= int(user_input) < len(directory_list):
            file_name = directory_list[int(user_input)]
            out("File: " + file_name)
            return file_name
        else:
            out("Invalid input: " + user_input)
            return get_file(path)
    except Exception:
        out("Invalid input: " + user_input)
        return get_file(path)


def get_worksheet(workbook):
    sheets = workbook.sheetnames
    print()
    for i in range(len(sheets)):
        print(" " + str(i) + " : " + sheets[i])
    print()
    user_input = ask("Enter number for results worksheet: ")
    try:
        if 0 <= int(user_input) < len(sheets):
            worksheet_name = sheets[int(user_input)]
            out("Worksheet: " + worksheet_name)
            return workbook.worksheets[int(user_input)]
        else:
            out("Invalid input: " + user_input)
            return get_worksheet(workbook)
    except Exception:
        out("Invalid input: " + user_input)
        return get_worksheet(workbook)


def get_survey_type():
    print()
    print(" 0: Transactional (NPS)")
    print(" 1: Annual")
    print()
    user_input = ask("Enter number for type of survey: ")
    try:
        if 0 <= int(user_input) <= 1:
            return int(user_input)
        else:
            out("Invalid input: " + user_input)
            return get_survey_type()
    except Exception:
        out("Invalid input: " + user_input)
        return get_survey_type()


def check_spreadsheet_columns(worksheet):
    config = configparser.ConfigParser()
    config.read('spreadsheet_config.ini')

    # check email column contains emails
    # looks for '@' in the first cell with a value
    config_column_email = config['mandatory_columns']['email']
    for row in range(2, worksheet.max_row + 1):
        cell = str(config_column_email) + str(row)
        value = worksheet[cell].value

        if value is None:
            continue

        if '@' not in value:
            out('Error: spreadsheet: email column ' + config_column_email + ' does not contain valid emails\n'
                + 'Cell value: ' + value + '\n'
                + 'Cell values in this column should contain the "@" symbol\n'
                + 'Please check that the email column is set correctly in the config file')
            sys.exit(1)

        else:  # '@' in value
            break

    # check end_date column contains dates
    # checks data type of first cell with a value
    config_column_end_date = config['mandatory_columns']['end_date']
    for row in range(2, worksheet.max_row + 1):
        cell = str(config_column_end_date) + str(row)
        value = worksheet[cell].value

        if value is None:
            continue

        if isinstance(value, str):
            if "AM" in value or "PM" in value:
                # format: MM/DD/YYYY HH:MM:SS AM/PM (081119 ITRS Client Survey Analysis Oct 19.xlsx)
                try:
                    value = value[:len(value) - 3]  # remove AM/PM
                    value = dt.strptime(value, '%m/%d/%Y %H:%M:%S')
                except:
                    out(
                        'Error: spreadsheet: end_date column ' + config_column_end_date + ' does not contain valid dates\n'
                        + 'Cell value: ' + value + '\n'
                        + 'Please check that the end_date column is set correctly in the config file')
                    sys.exit(1)

        if not isinstance(value, datetime.date):
            out('Error: spreadsheet: end_date column ' + config_column_end_date + ' does not contain valid dates\n'
                + 'Cell value: ' + value + '\n'
                + 'Please check that the end_date column is set correctly in the config file')
            sys.exit(1)

        else:
            break


def convert_score(score):
    if score == 0:
        return 0
    else:
        return (score * 2) + 2


################
# MAIN PROGRAM #
################

config = configparser.ConfigParser()
config.read('spreadsheet_config.ini')
check_config(config)

# DEVELOPMENT
# path = 'C:\\Users\\slowden\\ITRS Group Ltd\\ITRS Group Ltd Team Site - Intern\\Client Survey Processor\\Input Spreadsheets'
# file = '200729 client_survey annual survey week 1-4.xlsx'
# workbook = load_workbook(filename=os.path.join(path, file), read_only=True, data_only=True)
# worksheet = workbook.worksheets[0]
# PRODUCTION
path = get_path()
file = get_file(path)
workbook = load_workbook(filename=os.path.join(path, file), read_only=True, data_only=True)
worksheet = get_worksheet(workbook)

check_spreadsheet_columns(worksheet)

##############
# CSP_Survey #
##############
survey_id = db_insert_survey(file)

# process spreadsheet row by row
config_column_email = config['mandatory_columns']['email']
config_column_end_date = config['mandatory_columns']['end_date']
config_end_date_format = config['mandatory_columns']['end_date_format']
for row in range(2, worksheet.max_row + 1):

    ###################
    # CSP_Survey_User #
    ###################
    email = worksheet[str(config_column_email) + str(row)].value
    if email is None:
        out(f'Processing row [{str(row)} of {worksheet.max_row}] {email} - skipping')
        continue
    out(f'Processing row [{str(row)} of {worksheet.max_row}] {email}')
    user_id = sql_get_user_id(email)

    # date transformations
    end_date = worksheet[str(config_column_end_date) + str(row)].value
    if isinstance(end_date, str):
        if "AM" in end_date or "PM" in end_date:
            # MM/DD/YYYY HH:MM:SS AM/PM
            # 081119 ITRS Client Survey Analysis Oct 19.xlsx
            end_date = end_date[:len(end_date) - 3]  # remove AM/PM
            end_date = dt.strptime(end_date, '%m/%d/%Y %H:%M:%S')
        else:  # no seconds in end_date
            # DD/MM/YYYY HH:MM
            # 200723 client_survey annual survey week 1-3.xlsx
            end_date += ':00'  # append empty seconds
            end_date = dt.strptime(end_date, '%d/%m/%Y %H:%M:%S')
    else:
        if config_end_date_format == 'UK':
            # DD/MM/YYYY HH:MM:SS
            # 200729 client_survey annual survey week 1-4.xlsx
            end_date = dt.strptime(str(end_date), '%Y-%m-%d %H:%M:%S')
        else:  # config_end_date_format == 'US'
            # MM/DD/YYYY HH:MM:SS
            # 081119 ITRS Client Survey Analysis Oct 19.xlsx
            end_date = dt.strptime(str(end_date), '%Y-%d-%m %H:%M:%S')

    survey_user_id = db_insert_survey_user(survey_id, user_id, email, end_date)

    ################
    # CSP_Response #
    ################
    for i in range(1, sql_get_last_id('CSP_Question') + 1):
        if str(i) not in config:
            # Question i does not exist as section in config file
            continue

        config_score_max = config[str(i)]['score_max']
        config_column_score = config[str(i)]['score']
        config_column_comment = config[str(i)]['comment']

        if config_column_score == '0' and config_column_comment == '0':
            # Question i not included in survey
            continue
        elif config_column_score == '0':
            score = None
            comment = worksheet[str(config_column_comment) + str(row)].value
        elif config_column_comment == '0':
            score = worksheet[str(config_column_score) + str(row)].value
            comment = None
        else:
            score = worksheet[str(config_column_score) + str(row)].value
            comment = worksheet[str(config_column_comment) + str(row)].value

        if score is not None:
            try:
                score = int(score)
                if config_score_max == '4':
                    score = convert_score(score)
            except ValueError:
                score = None

        db_insert_response(survey_user_id, i, score, comment)

print()
out(f'{file} has been processed.')
